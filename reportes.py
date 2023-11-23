import json
import requests
import time
import os
import errno
from openpyxl import Workbook, load_workbook

def reescribirArchivos():
    escritura=[]
    nombredelarchivo=input("\nQue archivo quiere ver?(Solo Consultas): ")
    try:
        with open("Consulta/" + nombredelarchivo + ".txt","r") as archivo:
            print("\nArchivo leido correctamente.")
            texto=input("\nPor favor escriba lo que quiere agregar al reporte: ")
            wb=Workbook()
            nDAs=str("_Reporte_" + "_" + time.strftime("%d") + "_" + time.strftime("%m") + "_" + time.strftime("%Y") + "_" + time.strftime("%H") + "_" + time.strftime("%M"))
            filesheet = "Reportes/"+ nombredelarchivo + nDAs + ".xlsx"
            wb.save(filesheet)
            wb = load_workbook(filesheet)
            sheet = wb.active
            escritura.append(("ID","Categoria","Nombre","Precio","Etiquetas"))
            for linea in archivo:
                principio=linea.find("ID:")
                fin=linea.find(",\tCategoria:")
                i=(linea[principio+3:fin])
                
                principio=linea.find(",\tCategoria:")
                fin=linea.find(",\tNombre:")
                cate=(linea[principio+13:fin])
                
                principio=linea.find("Nombre:")
                fin=linea.find(",\tPrecio:")
                nom=(linea[principio+8:fin])
                
                principio=linea.find("Precio:")
                fin=linea.find(",\tEtiquetas:")
                pre=(linea[principio+7:fin])
                
                principio=linea.find("Etiquetas:")
                eti=(linea[principio+10:len(linea)])
                
                escritura.append((i,cate,nom,pre,eti))
            escritura.append(("Reporte: "+texto," "," "," "," "))
            thisescritura=tuple(escritura)
            for row in escritura:
                sheet.append(row)
            wb.save(filesheet)
            print("Se Exporto Exitosamente el Archivo Excel llamado "+ nombredelarchivo + nDAs + ".")
    except FileNotFoundError:
        print("\nError: El archivo con el nombre " + nombredelarchivo + " no se encontró.\n")






def leerReportes():
    nombredelarchivo=input("\nQue archivo quiere ver?(Solo Reportes): ")
    filesheet = "Reportes/"+ nombredelarchivo + ".xlsx"
    wb = load_workbook(filesheet)
    sheer = wb.active
    for line in range(sheer.max_row):
        cell0=sheer.cell(row=line+1, column = 1)
        cell1=sheer.cell(row=line+1, column = 2)
        cell2=sheer.cell(row=line+1, column = 3)
        cell3=sheer.cell(row=line+1, column = 4)
        cell4=sheer.cell(row=line+1, column = 5)
        print(str(cell0.value)+ "\t" + str(cell1.value) + "\t" + str(cell2.value) + "\t" + str(cell3.value) + "\t" + str(cell4.value))
    wb.save(filesheet)